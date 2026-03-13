"""
Sistema de Gestión de Asistencia Escolar - MEC Paraguay
Versión 2.1 — Login, botones P/A/J estilo card, exportación Excel,
importación PDF, edición de estudiantes, logo institucional.
"""

import streamlit as st
import psycopg2
import psycopg2.extras
import pandas as pd
import plotly.express as px
from datetime import date, timedelta
import random
import io
import base64
import json
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIGURACIÓN GENERAL
# ─────────────────────────────────────────────
UMBRAL_FALTAS_CONSECUTIVAS = 2

GRADOS = {
    "Tercer Ciclo": ["7° Grado", "8° Grado", "9° Grado"],
    "Nivel Medio - BTS": ["1° BTS"],
    "Nivel Medio - BC":  ["1° BC", "2° BC", "3° BC"],
    "Nivel Medio - BTI": ["1° BTI", "2° BTI", "3° BTI"],
}
TODOS_LOS_GRADOS = [g for nivel in GRADOS.values() for g in nivel]
ESTADOS = ["Presente", "Ausente Injustificado", "Ausente Justificado"]
TURNOS  = ["Mañana", "Tarde"]

# ─────────────────────────────────────────────
# LOGIN
# ─────────────────────────────────────────────

# Token de sesión — se guarda en la URL para sobrevivir recargas
import hashlib, os

def _generar_token():
    """Genera un token único por sesión basado en secrets."""
    base = st.secrets.get("app_usuario", "Lucasmen") + st.secrets.get("app_password", "123456")
    return hashlib.sha256(base.encode()).hexdigest()[:16]


def pagina_login():
    """Pantalla de inicio de sesión."""
    st.markdown("""
    <style>
    [data-testid="stSidebar"] { display: none !important; }
    [data-testid="stHeader"]  { display: none !important; }
    .main .block-container {
        max-width: 440px !important;
        margin: 0 auto !important;
        padding-top: 80px !important;
    }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div style="text-align:center;margin-bottom:32px;">
        <div style="font-size:56px;">🏫</div>
        <div style="font-size:26px;font-weight:700;margin-top:8px;">Asistencia Escolar</div>
        <div style="font-size:13px;opacity:0.6;margin-top:4px;">MEC Paraguay &middot; Tercer Ciclo &amp; Nivel Medio</div>
    </div>
    """, unsafe_allow_html=True)

    usuario    = st.text_input("👤 Usuario", placeholder="Ingresá tu usuario")
    contrasena = st.text_input("🔒 Contraseña", type="password", placeholder="Ingresá tu contraseña")
    st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

    if st.button("Ingresar", type="primary", use_container_width=True):
        usuario_ok = st.secrets.get("app_usuario", "Lucasmen")
        clave_ok   = st.secrets.get("app_password", "123456")
        if usuario == usuario_ok and contrasena == clave_ok:
            st.session_state["autenticado"] = True
            # Guardar token en URL — sobrevive recargas y deslizar en tablet
            st.query_params["t"] = _generar_token()
            st.rerun()
        else:
            st.error("❌ Usuario o contraseña incorrectos.")


def verificar_login():
    """Devuelve True si autenticado via session_state o token en URL."""
    if st.session_state.get("autenticado"):
        return True
    # Verificar token en query params (persiste entre recargas)
    token_url = st.query_params.get("t", "")
    if token_url and token_url == _generar_token():
        st.session_state["autenticado"] = True
        return True
    return False


# ─────────────────────────────────────────────
# CONEXIÓN BASE DE DATOS
# ─────────────────────────────────────────────

def _nueva_conexion():
    return psycopg2.connect(
        host=st.secrets["db_host"],
        port=st.secrets["db_port"],
        dbname=st.secrets["db_name"],
        user=st.secrets["db_user"],
        password=st.secrets["db_password"],
        sslmode="require",
        connect_timeout=10,
    )


def get_conn():
    """Devuelve conexión activa. Si está caída o en error, reconecta automáticamente."""
    if "db_conn" not in st.session_state or st.session_state["db_conn"] is None:
        st.session_state["db_conn"] = _nueva_conexion()
    conn = st.session_state["db_conn"]
    try:
        if conn.closed:
            raise Exception("cerrada")
        if conn.status == psycopg2.extensions.STATUS_IN_TRANSACTION:
            conn.rollback()
    except Exception:
        try:
            conn.close()
        except Exception:
            pass
        conn = _nueva_conexion()
        st.session_state["db_conn"] = conn
    return conn


def _forzar_nueva_conexion():
    """Cierra y elimina la conexión actual para forzar una nueva."""
    try:
        if "db_conn" in st.session_state:
            st.session_state["db_conn"].close()
    except Exception:
        pass
    st.session_state.pop("db_conn", None)


def run_query(sql, params=None, fetch=True):
    for intento in range(3):
        try:
            conn = get_conn()
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(sql, params or ())
                if fetch:
                    return cur.fetchall()
                conn.commit()
                return None
        except Exception as e:
            _forzar_nueva_conexion()
            if intento == 2:
                raise e
    return None


def run_df(sql, params=None) -> pd.DataFrame:
    rows = run_query(sql, params, fetch=True)
    return pd.DataFrame(rows) if rows else pd.DataFrame()


# ─────────────────────────────────────────────
# INIT BASE DE DATOS
# ─────────────────────────────────────────────

def init_db():
    conn = get_conn()
    with conn.cursor() as cur:
        # Check rápido: si ya está migrado completamente, salir inmediatamente
        try:
            cur.execute("SELECT valor FROM config WHERE clave='schema_version'")
            row = cur.fetchone()
            if row and row[0] == "3":
                return  # Ya migrado — salida instantánea
        except Exception:
            pass

        # Primera vez o migración pendiente — hacer todo
        cur.execute("""
            CREATE TABLE IF NOT EXISTS config (clave TEXT PRIMARY KEY, valor TEXT);
            CREATE TABLE IF NOT EXISTS grados (
                id SERIAL PRIMARY KEY, nombre TEXT UNIQUE NOT NULL, nivel TEXT NOT NULL);
            CREATE TABLE IF NOT EXISTS estudiantes (
                id SERIAL PRIMARY KEY, nombre TEXT NOT NULL, ci TEXT,
                grado_id INTEGER NOT NULL REFERENCES grados(id), contacto TEXT);
            CREATE TABLE IF NOT EXISTS asistencia (
                id SERIAL PRIMARY KEY,
                estudiante_id INTEGER NOT NULL REFERENCES estudiantes(id),
                fecha DATE NOT NULL, turno TEXT NOT NULL DEFAULT 'Mañana',
                estado TEXT NOT NULL CHECK(estado IN (
                    'Presente','Ausente Injustificado','Ausente Justificado')),
                UNIQUE(estudiante_id, fecha, turno));
        """)
        cur.execute("ALTER TABLE estudiantes ADD COLUMN IF NOT EXISTS ci TEXT;")
        cur.execute("ALTER TABLE asistencia ADD COLUMN IF NOT EXISTS turno TEXT DEFAULT 'Mañana';")
        cur.execute("UPDATE asistencia SET turno='Mañana' WHERE turno IS NULL;")
        cur.execute("""
            DO $do$ BEGIN
                IF EXISTS (SELECT 1 FROM pg_constraint WHERE conname='asistencia_estudiante_id_fecha_key')
                THEN ALTER TABLE asistencia DROP CONSTRAINT asistencia_estudiante_id_fecha_key; END IF;
                IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname='asistencia_estudiante_id_fecha_turno_key')
                THEN ALTER TABLE asistencia ADD CONSTRAINT asistencia_estudiante_id_fecha_turno_key
                    UNIQUE(estudiante_id, fecha, turno); END IF;
            END $do$;
        """)
        cur.execute("UPDATE grados SET nombre=REPLACE(nombre,'BTC','BC'), nivel=REPLACE(nivel,'BTC','BC') WHERE nombre LIKE '%%BTC%%'")
        for nivel, lista in GRADOS.items():
            for grado in lista:
                cur.execute("INSERT INTO grados (nombre,nivel) VALUES (%s,%s) ON CONFLICT (nombre) DO UPDATE SET nivel=EXCLUDED.nivel", (grado, nivel))
        for clave, valor in [("institucion_nombre","Institución Educativa"),("institucion_logo","")]:
            cur.execute("INSERT INTO config (clave,valor) VALUES (%s,%s) ON CONFLICT (clave) DO NOTHING", (clave, valor))
        # Marcar schema como completamente migrado
        cur.execute("INSERT INTO config (clave,valor) VALUES ('schema_version','3') ON CONFLICT (clave) DO UPDATE SET valor='3'")
        conn.commit()


def seed_mock_data():
    conn = get_conn()
    with conn.cursor() as cur:
        cur.execute("SELECT COUNT(*) FROM estudiantes")
        if cur.fetchone()[0] > 0:
            return
        mock = {
            "7° Grado": [("Ana Martínez", "1234567"), ("Luis Pérez", "2345678"),
                         ("Sofía Rojas", "3456789"), ("Carlos López", "4567890")],
            "8° Grado": [("Miguel Torres", "5678901"), ("Lucía Ramírez", "6789012"),
                         ("Diego Flores", "7890123")],
            "1° BTS":   [("Fernando Ríos", "8901234"), ("Camila Díaz", "9012345")],
        }
        today = date.today()
        for grado_nombre, alumnos in mock.items():
            cur.execute("SELECT id FROM grados WHERE nombre = %s", (grado_nombre,))
            row = cur.fetchone()
            if not row:
                continue
            grado_id = row[0]
            for nombre, ci in alumnos:
                cur.execute(
                    "INSERT INTO estudiantes (nombre, ci, grado_id) VALUES (%s, %s, %s) RETURNING id",
                    (nombre, ci, grado_id),
                )
                est_id = cur.fetchone()[0]
                dias_sim, dia_actual = 0, today
                while dias_sim < 5:
                    if dia_actual.weekday() < 5:
                        estado = random.choices(ESTADOS, weights=[0.80, 0.12, 0.08], k=1)[0]
                        cur.execute(
                            "INSERT INTO asistencia (estudiante_id, fecha, estado) VALUES (%s,%s,%s) ON CONFLICT DO NOTHING",
                            (est_id, dia_actual, estado),
                        )
                        dias_sim += 1
                    dia_actual -= timedelta(days=1)
        conn.commit()


# ─────────────────────────────────────────────
# CONFIG INSTITUCIONAL
# ─────────────────────────────────────────────

def get_config(clave):
    rows = run_query("SELECT valor FROM config WHERE clave = %s", (clave,))
    return rows[0]["valor"] if rows else ""


def set_config(clave, valor):
    run_query(
        "INSERT INTO config (clave,valor) VALUES (%s,%s) ON CONFLICT (clave) DO UPDATE SET valor=EXCLUDED.valor",
        (clave, valor), fetch=False,
    )


# ─────────────────────────────────────────────
# FUNCIONES DE DATOS
# ─────────────────────────────────────────────

def get_estudiantes_por_grado(grado_nombre):
    ck = f"estxgrado_{grado_nombre}"
    if ck not in st.session_state:
        st.session_state[ck] = run_df("""
            SELECT e.id, e.nombre, e.ci, e.contacto
            FROM estudiantes e JOIN grados g ON e.grado_id = g.id
            WHERE g.nombre = %s ORDER BY e.nombre
        """, (grado_nombre,))
    return st.session_state[ck]


def get_asistencia_fecha(grado_nombre, fecha, turno="Mañana"):
    cache_k = f"asist_{grado_nombre}_{fecha}_{turno}"
    if cache_k not in st.session_state:
        st.session_state[cache_k] = run_df("""
            SELECT e.id as estudiante_id, e.nombre, e.ci,
                   COALESCE(a.estado, 'Sin registro') as estado
            FROM estudiantes e JOIN grados g ON e.grado_id = g.id
            LEFT JOIN asistencia a ON a.estudiante_id = e.id
                AND a.fecha = %s AND a.turno = %s
            WHERE g.nombre = %s ORDER BY e.nombre
        """, (fecha, turno, grado_nombre))
    return st.session_state[cache_k]


def guardar_un_estado(est_id, fecha, turno, estado):
    """Guarda un solo registro inmediatamente — llamado al tocar P/A/J."""
    run_query("""
        INSERT INTO asistencia (estudiante_id, fecha, turno, estado)
        VALUES (%s, %s, %s, %s)
        ON CONFLICT (estudiante_id, fecha, turno)
        DO UPDATE SET estado = EXCLUDED.estado
    """, (int(est_id), fecha, turno, estado), fetch=False)


def autoguardar_lista(df, fecha, turno, sk_fn):
    """Guarda silenciosamente toda la lista actual desde session_state."""
    OPCION_A_ESTADO = {"P": "Presente", "A": "Ausente Injustificado", "J": "Ausente Justificado"}
    if df.empty:
        return
    registros = []
    for _, row in df.iterrows():
        eid    = int(row["estudiante_id"])
        opcion = st.session_state.get(sk_fn(eid), "P")
        registros.append((eid, fecha, turno, OPCION_A_ESTADO.get(opcion, "Presente")))
    try:
        guardar_asistencia(registros, turno=turno)
    except Exception:
        pass  # Silencioso — no interrumpir al usuario


def guardar_asistencia(registros, turno="Mañana"):
    conn = get_conn()
    with conn.cursor() as cur:
        for est_id, fecha, estado in registros:
            cur.execute("""
                INSERT INTO asistencia (estudiante_id, fecha, turno, estado) VALUES (%s,%s,%s,%s)
                ON CONFLICT (estudiante_id, fecha, turno) DO UPDATE SET estado=EXCLUDED.estado
            """, (est_id, fecha, turno, estado))
        conn.commit()


def get_resumen_grado(grado_nombre):
    ck = f"resumen_{grado_nombre}"
    if ck not in st.session_state:
        st.session_state[ck] = run_df("""
            SELECT e.nombre,
                   COUNT(CASE WHEN a.estado='Presente' THEN 1 END) as presentes,
                   COUNT(CASE WHEN a.estado='Ausente Injustificado' THEN 1 END) as inj,
                   COUNT(CASE WHEN a.estado='Ausente Justificado' THEN 1 END) as just,
                   COUNT(a.id) as total_dias
            FROM estudiantes e JOIN grados g ON e.grado_id=g.id
            LEFT JOIN asistencia a ON a.estudiante_id=e.id
            WHERE g.nombre=%s GROUP BY e.id, e.nombre ORDER BY e.nombre
        """, (grado_nombre,))
    return st.session_state[ck]


def get_asistencia_rango(grado_nombre, fecha_ini, fecha_fin):
    return run_df("""
        SELECT e.nombre, e.ci, a.fecha, a.turno, a.estado, g.nombre as grado
        FROM asistencia a
        JOIN estudiantes e ON a.estudiante_id=e.id
        JOIN grados g ON e.grado_id=g.id
        WHERE g.nombre=%s AND a.fecha BETWEEN %s AND %s
        ORDER BY e.nombre, a.fecha, a.turno
    """, (grado_nombre, fecha_ini, fecha_fin))


def get_reporte_estudiante(est_id):
    ck = f"rep_{est_id}"
    if ck not in st.session_state:
        st.session_state[ck] = run_df("""
            SELECT fecha, turno, estado
            FROM asistencia
            WHERE estudiante_id = %s
            ORDER BY fecha DESC, turno
            LIMIT 60
        """, (int(est_id),))
    return st.session_state[ck]


def detectar_faltas_consecutivas(grado_nombre=None):
    ck = f"alertas_{grado_nombre}"
    if ck in st.session_state:
        return st.session_state[ck]
    filtro = "AND g.nombre = %s" if grado_nombre else ""
    params = (grado_nombre,) if grado_nombre else ()
    df = run_df(f"""
        SELECT e.id as estudiante_id, e.nombre, e.contacto, g.nombre as grado,
               a.fecha, a.estado
        FROM asistencia a
        JOIN estudiantes e ON a.estudiante_id=e.id
        JOIN grados g ON e.grado_id=g.id
        WHERE a.estado LIKE 'Ausente%%' {filtro}
        ORDER BY e.id, a.fecha DESC
    """, params)

    if df.empty:
        empty = pd.DataFrame(columns=["nombre","grado","contacto","faltas_consecutivas","desde"])
        st.session_state[ck] = empty
        return empty

    resultados = []
    for est_id, grupo in df.groupby("estudiante_id"):
        fechas = sorted(pd.to_datetime(grupo["fecha"]).tolist(), reverse=True)
        racha = 1
        for i in range(1, len(fechas)):
            diff = (fechas[i-1] - fechas[i]).days
            if 1 <= diff <= 3:
                racha += 1
            else:
                break
        if racha >= UMBRAL_FALTAS_CONSECUTIVAS:
            resultados.append({
                "nombre": grupo["nombre"].iloc[0],
                "grado": grupo["grado"].iloc[0],
                "contacto": grupo["contacto"].iloc[0] or "Sin contacto",
                "faltas_consecutivas": racha,
                "desde": fechas[-1].date() if racha > 1 else fechas[0].date(),
            })
    result = pd.DataFrame(resultados).sort_values("faltas_consecutivas", ascending=False) if resultados else pd.DataFrame(columns=["nombre","grado","contacto","faltas_consecutivas","desde"])
    st.session_state[ck] = result
    return result


def agregar_estudiante(nombre, ci, grado_nombre, contacto):
    nombre   = (nombre   or "").strip()
    ci       = (ci       or "").strip()
    contacto = (contacto or "").strip()
    rows = run_query("SELECT id FROM grados WHERE nombre=%s", (grado_nombre,))
    if rows:
        grado_id = rows[0]["id"]
        run_query(
            "INSERT INTO estudiantes (nombre,ci,grado_id,contacto) VALUES (%s,%s,%s,%s)",
            (nombre, ci, grado_id, contacto),
            fetch=False,
        )
        for k in [k for k in st.session_state if k.startswith(("asist_","estxgrado_","resumen_"))]:
            st.session_state.pop(k, None)


def actualizar_estudiante(est_id, nombre, ci, grado_nombre, contacto):
    est_id   = int(est_id)   # convertir numpy.int64 → int nativo
    nombre   = (nombre   or "").strip()
    ci       = (ci       or "").strip()
    contacto = (contacto or "").strip()
    _forzar_nueva_conexion()
    rows = run_query("SELECT id FROM grados WHERE nombre=%s", (grado_nombre,))
    if rows:
        grado_id = int(rows[0]["id"])
        run_query(
            "UPDATE estudiantes SET nombre=%s,ci=%s,grado_id=%s,contacto=%s WHERE id=%s",
            (nombre, ci, grado_id, contacto, est_id),
            fetch=False,
        )
        for k in [k for k in st.session_state if k.startswith(("asist_","estxgrado_","resumen_"))]:
            st.session_state.pop(k, None)


def eliminar_estudiante(est_id):
    est_id = int(est_id)
    run_query("DELETE FROM asistencia WHERE estudiante_id=%s", (est_id,), fetch=False)
    run_query("DELETE FROM estudiantes WHERE id=%s", (est_id,), fetch=False)
    for k in [k for k in st.session_state if k.startswith(("asist_","estxgrado_","resumen_"))]:
        st.session_state.pop(k, None)


# ─────────────────────────────────────────────
# IMPORTACIÓN PDF
# ─────────────────────────────────────────────

def extraer_alumnos_pdf(pdf_bytes):
    filas = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for pagina in pdf.pages:
            for tabla in (pagina.extract_tables() or []):
                if not tabla:
                    continue
                header = [str(c).lower().strip() if c else "" for c in tabla[0]]
                col_nombre = next((i for i, h in enumerate(header) if "nombre" in h or "apellido" in h), 0)
                col_ci = next((i for i, h in enumerate(header) if "ci" in h or "cédula" in h or "cedula" in h), 1)
                for fila in tabla[1:]:
                    if not fila or not fila[col_nombre]:
                        continue
                    nombre = str(fila[col_nombre]).strip()
                    ci = str(fila[col_ci]).strip() if col_ci < len(fila) and fila[col_ci] else ""
                    if nombre and nombre.lower() not in ("nombre","apellido","alumno",""):
                        filas.append({"nombre": nombre, "ci": ci})
    return pd.DataFrame(filas)


# ─────────────────────────────────────────────
# EXPORTACIÓN EXCEL
# ─────────────────────────────────────────────

def _hacer_hoja_diaria(wb, titulo, df_turno, alumnos, fechas, hfill, hfont, brd, tfont, verde, rojo, amarillo, grado_nombre, fecha_ini, fecha_fin, institucion):
    """Crea una hoja con filas=alumnos, columnas=fechas, valores=P/A/J."""
    ws = wb.create_sheet(titulo)
    ESTADO_CORTO = {"Presente": "P", "Ausente Injustificado": "A", "Ausente Justificado": "J"}
    COLOR_PAJ    = {"P": "C6EFCE", "A": "FFC7CE", "J": "FFEB9C"}

    # Título
    ncols = 2 + len(fechas)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(1,1, f"{institucion} — {grado_nombre} — {titulo}").font = tfont
    ws.cell(1,1).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(2,1, f"{fecha_ini.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}").font = Font(italic=True, size=10)
    ws.cell(2,1).alignment = Alignment(horizontal="center")

    # Encabezados
    for ci, h in enumerate(["Nombre", "CI"], 1):
        c = ws.cell(3, ci, h); c.fill=hfill; c.font=hfont
        c.alignment=Alignment(horizontal="center"); c.border=brd

    DIAS_ES = ["Lun","Mar","Mié","Jue","Vie","Sáb","Dom"]
    for fi, fecha in enumerate(fechas):
        lbl = DIAS_ES[fecha.weekday()] + "\n" + fecha.strftime('%d/%m')
        c = ws.cell(3, 3+fi, lbl); c.fill=hfill; c.font=hfont
        c.alignment=Alignment(horizontal="center", wrap_text=True); c.border=brd
    ws.row_dimensions[3].height = 30

    # Construir lookup: (nombre, fecha) → estado corto
    lookup = {}
    for _, r in df_turno.iterrows():
        f = r["fecha"] if isinstance(r["fecha"], date) else pd.to_datetime(r["fecha"]).date()
        lookup[(r["nombre"], f)] = ESTADO_CORTO.get(r["estado"], "")

    # Filas de alumnos
    for ri, (nombre, ci_val) in enumerate(alumnos):
        row_n = 4 + ri
        c = ws.cell(row_n, 1, nombre); c.border=brd; c.alignment=Alignment(horizontal="left")
        c = ws.cell(row_n, 2, ci_val); c.border=brd; c.alignment=Alignment(horizontal="center")
        for fi, fecha in enumerate(fechas):
            val = lookup.get((nombre, fecha), "")
            c = ws.cell(row_n, 3+fi, val)
            c.border=brd; c.alignment=Alignment(horizontal="center")
            if val in COLOR_PAJ:
                c.fill = PatternFill("solid", fgColor=COLOR_PAJ[val])

    # Anchos de columna
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 12
    for fi in range(len(fechas)):
        ws.column_dimensions[get_column_letter(3+fi)].width = 8


def generar_excel_resumen(grado_nombre, fecha_ini, fecha_fin, institucion):
    df_todo = get_asistencia_rango(grado_nombre, fecha_ini, fecha_fin)

    wb   = openpyxl.Workbook()
    hfill    = PatternFill("solid", fgColor="1F4E79")
    hfont    = Font(color="FFFFFF", bold=True, size=11)
    tfont    = Font(bold=True, size=14, color="1F4E79")
    brd      = Border(left=Side(style="thin"), right=Side(style="thin"),
                      top=Side(style="thin"), bottom=Side(style="thin"))
    verde    = PatternFill("solid", fgColor="C6EFCE")
    rojo     = PatternFill("solid", fgColor="FFC7CE")
    amarillo = PatternFill("solid", fgColor="FFEB9C")

    # ── Hoja 1: Resumen General ──
    resumen = get_resumen_grado(grado_nombre)
    ws1 = wb.active
    ws1.title = "Resumen General"

    ws1.merge_cells("A1:G1")
    ws1["A1"] = f"{institucion} — Registro de Asistencia"
    ws1["A1"].font = tfont
    ws1["A1"].alignment = Alignment(horizontal="center")
    ws1.merge_cells("A2:G2")
    ws1["A2"] = f"{grado_nombre}   |   {fecha_ini.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}"
    ws1["A2"].alignment = Alignment(horizontal="center")
    ws1["A2"].font = Font(italic=True, size=10)

    for ci, h in enumerate(["Nombre","CI","Presentes","F. Injustificadas","F. Justificadas","Días","% Asistencia"], 1):
        c = ws1.cell(row=4, column=ci, value=h)
        c.fill=hfill; c.font=hfont; c.alignment=Alignment(horizontal="center"); c.border=brd

    est_ci = run_df("SELECT e.nombre,e.ci FROM estudiantes e JOIN grados g ON e.grado_id=g.id WHERE g.nombre=%s ORDER BY e.nombre", (grado_nombre,))
    ci_map  = dict(zip(est_ci["nombre"], est_ci["ci"])) if not est_ci.empty else {}
    alumnos = [(r["nombre"], ci_map.get(r["nombre"],"")) for _, r in est_ci.iterrows()]

    for ri, row in resumen.iterrows():
        pct = round(row["presentes"]/row["total_dias"]*100, 1) if row["total_dias"] > 0 else 0
        for ci, val in enumerate([row["nombre"], ci_map.get(row["nombre"],""),
                                   row["presentes"], row["inj"], row["just"],
                                   row["total_dias"], f"{pct}%"], 1):
            c = ws1.cell(row=ri+4, column=ci, value=val)
            c.border=brd; c.alignment=Alignment(horizontal="left" if ci==1 else "center")
            if ci==7: c.fill = verde if pct>=75 else rojo
    for i, w in enumerate([30,12,12,18,16,12,14], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Fechas únicas del rango ──
    if not df_todo.empty:
        fechas_unicas = sorted(set(
            r if isinstance(r, date) else pd.to_datetime(r).date()
            for r in df_todo["fecha"]
        ))
        turnos_presentes = sorted(df_todo["turno"].dropna().unique())

        # ── Hojas por turno: filas=alumnos, columnas=fechas ──
        for turno in turnos_presentes:
            df_t = df_todo[df_todo["turno"] == turno]
            _hacer_hoja_diaria(
                wb, turno, df_t, alumnos, fechas_unicas,
                hfill, hfont, brd, tfont, verde, rojo, amarillo,
                grado_nombre, fecha_ini, fecha_fin, institucion
            )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ─────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────

def inject_css():
    st.markdown("""
    <style>
    html, body, [class*="css"] { font-size: 16px !important; }

    /* ── Card alumno ── */
    .alumno-card {
        background: #1e2130;
        border-radius: 14px;
        padding: 14px 18px;
        margin-bottom: 10px;
        display: flex;
        align-items: center;
        justify-content: space-between;
        box-shadow: 0 2px 8px rgba(0,0,0,0.18);
    }
    .alumno-info { flex: 1; }
    .alumno-nombre {
        font-size: 15px;
        font-weight: 700;
        color: #0f1117;
        text-transform: uppercase;
        letter-spacing: 0.3px;
    }
    .alumno-ci { font-size: 12px; color: #555; margin-top: 2px; }

    /* ── Botones P/A/J circulares ── */
    .paj-btns {
        display: flex;
        gap: 10px;
        align-items: center;
    }
    .paj-btn {
        width: 44px; height: 44px;
        border-radius: 50%;
        border: 2px solid;
        font-weight: 700;
        font-size: 16px;
        cursor: pointer;
        display: flex; align-items: center; justify-content: center;
        transition: all 0.15s;
        flex-shrink: 0;
    }
    .paj-P-off { border-color:#2ecc71; color:#2ecc71; background:transparent; }
    .paj-P-on  { border-color:#2ecc71; color:#fff;    background:#2ecc71; }
    .paj-A-off { border-color:#e74c3c; color:#e74c3c; background:transparent; }
    .paj-A-on  { border-color:#e74c3c; color:#fff;    background:#e74c3c; }
    .paj-J-off { border-color:#f39c12; color:#f39c12; background:transparent; }
    .paj-J-on  { border-color:#f39c12; color:#fff;    background:#f39c12; }

    /* ── Botones streamlit generales ── */
    .stButton > button {
        min-height: 48px !important;
        font-size: 15px !important;
        border-radius: 10px !important;
    }
    .stSelectbox > div > div,
    .stDateInput > div > div > input {
        min-height: 48px !important;
        font-size: 15px !important;
        border-radius: 8px !important;
    }
    .stTextInput > div > div > input {
        min-height: 48px !important;
        font-size: 15px !important;
        border-radius: 8px !important;
    }
    .stRadio label { font-size: 15px !important; padding: 8px 10px !important; border-radius: 8px !important; }
    .stTabs [data-baseweb="tab"] { min-height: 46px !important; font-size: 14px !important; }
    [data-testid="metric-container"] [data-testid="stMetricValue"] { font-size: 26px !important; }
    [data-testid="stSidebar"] { min-width: 240px !important; }
    .main .block-container { padding-top: 1.2rem !important; padding-bottom: 3rem !important; max-width: 1100px; }
    </style>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────
# PÁGINAS
# ─────────────────────────────────────────────

def pagina_pasar_lista():
    st.header("📋 Pasar Lista")

    col1, col2, col3 = st.columns([3, 2, 1])
    with col1:
        grado_sel = st.selectbox("Grado", TODOS_LOS_GRADOS, key="lista_grado")
    with col2:
        turno_sel = st.radio("Turno", TURNOS, horizontal=True, key="lista_turno")
    with col3:
        fecha_sel = st.date_input("Fecha", value=date.today(), key="lista_fecha")

    ESTADO_A_OPCION = {"Presente": "P", "Ausente Injustificado": "A", "Ausente Justificado": "J"}
    OPCION_A_ESTADO = {"P": "Presente", "A": "Ausente Injustificado", "J": "Ausente Justificado"}

    def sk(eid): return f"est_{grado_sel}_{fecha_sel}_{turno_sel}_{eid}"

    df = get_asistencia_fecha(grado_sel, fecha_sel, turno_sel)

    # ── Autoguardar lista anterior si cambió grado/fecha/turno ──
    prev_key = st.session_state.get("lista_prev_key")
    cur_key  = f"{grado_sel}|{fecha_sel}|{turno_sel}"
    if prev_key and prev_key != cur_key:
        # Recuperar df del grado anterior y guardarlo
        prev_parts = prev_key.split("|")
        if len(prev_parts) == 3:
            prev_grado, prev_fecha_str, prev_turno = prev_parts
            try:
                prev_fecha = date.fromisoformat(prev_fecha_str)
                prev_ck    = f"asist_{prev_grado}_{prev_fecha_str}_{prev_turno}"
                prev_df    = st.session_state.get(prev_ck)
                if prev_df is not None and not prev_df.empty:
                    def prev_sk(eid): return f"est_{prev_grado}_{prev_fecha_str}_{prev_turno}_{eid}"
                    autoguardar_lista(prev_df, prev_fecha, prev_turno, prev_sk)
            except Exception:
                pass
    st.session_state["lista_prev_key"] = cur_key

    # Inicializar estados desde BD solo una vez por grado+fecha+turno
    cache_key   = f"cache_{grado_sel}_{fecha_sel}_{turno_sel}"
    grados_init = st.session_state.setdefault("grados_init", set())
    if cache_key not in grados_init and not df.empty:
        for _, row in df.iterrows():
            eid = int(row["estudiante_id"])
            if sk(eid) not in st.session_state:
                e = row["estado"] if row["estado"] in ESTADOS else "Presente"
                st.session_state[sk(eid)] = ESTADO_A_OPCION.get(e, "P")
        grados_init.add(cache_key)

    if df.empty:
        st.warning(f"⚠️ No hay estudiantes en **{grado_sel}**.")
        return

    # ── Cabecera ──
    col_info, col_agregar = st.columns([3, 1])
    with col_info:
        total = len(df)
        ausentes_n = sum(1 for _, r in df.iterrows() if st.session_state.get(sk(int(r["estudiante_id"])), "P") == "A")
        just_n     = sum(1 for _, r in df.iterrows() if st.session_state.get(sk(int(r["estudiante_id"])), "P") == "J")
        st.markdown(f"**{total}** alumnos · 🔴 {ausentes_n} ausentes · 🟡 {just_n} justificados")
    with col_agregar:
        if st.button("➕ Agregar", use_container_width=True):
            st.session_state["mostrar_form_agregar_lista"] = not st.session_state.get("mostrar_form_agregar_lista", False)
            st.rerun()

    if st.session_state.get("mostrar_form_agregar_lista"):
        with st.form("form_agregar_rapido", clear_on_submit=True):
            st.markdown(f"**Agregar a {grado_sel}**")
            c1, c2 = st.columns(2)
            nuevo_nombre = c1.text_input("Nombre", placeholder="APELLIDO, Nombre")
            nuevo_ci     = c2.text_input("CI", placeholder="ej: 5123456")
            if st.form_submit_button("✅ Agregar", type="primary", use_container_width=True):
                if nuevo_nombre.strip():
                    try:
                        agregar_estudiante(nuevo_nombre.strip().upper(), nuevo_ci.strip(), grado_sel, "")
                        st.session_state.pop(f"asist_{grado_sel}_{fecha_sel}_{turno_sel}", None)
                        grados_init.discard(cache_key)
                        st.session_state["mostrar_form_agregar_lista"] = False
                        st.rerun()
                    except Exception as ex:
                        st.error(f"❌ Error: {ex}")
                else:
                    st.warning("⚠️ Ingresá el nombre.")

    ca, cb = st.columns(2)
    with ca:
        if st.button("✅ Todos Presentes", use_container_width=True):
            for _, row in df.iterrows():
                st.session_state[sk(int(row["estudiante_id"]))] = "P"
            st.rerun()
    with cb:
        if st.button("💾 Guardar todo", type="primary", use_container_width=True):
            registros = [(int(r["estudiante_id"]), fecha_sel,
                         OPCION_A_ESTADO.get(st.session_state.get(sk(int(r["estudiante_id"])), "P"), "Presente"))
                         for _, r in df.iterrows()]
            try:
                guardar_asistencia(registros, turno=turno_sel)
                st.session_state.pop(f"asist_{grado_sel}_{fecha_sel}_{turno_sel}", None)
                st.session_state.pop("metrics_ts", None)
                st.success(f"✅ Lista completa guardada — {grado_sel} {turno_sel} {fecha_sel.strftime('%d/%m/%Y')}")
                st.balloons()
            except Exception as ex:
                st.error(f"❌ {ex}")

    st.markdown("---")

    # ── Lista alumnos ──
    for _, row in df.iterrows():
        eid    = int(row["estudiante_id"])
        nombre = str(row["nombre"])
        ci     = str(row.get("ci", "") or "")
        estado = st.session_state.get(sk(eid), "P")

        lbl_p = "🟢 P" if estado == "P" else "P"
        lbl_a = "🔴 A" if estado == "A" else "A"
        lbl_j = "🟡 J" if estado == "J" else "J"

        c_nom, c_p, c_a, c_j, c_rep = st.columns([5, 1, 1, 1, 1])
        with c_nom:
            st.markdown(f"**{nombre}**")
            if ci:
                st.caption(ci)
        with c_p:
            if st.button(lbl_p, key=f"p_{eid}", use_container_width=True):
                st.session_state[sk(eid)] = "P"
                guardar_un_estado(eid, fecha_sel, turno_sel, "Presente")
                st.session_state.pop(f"asist_{grado_sel}_{fecha_sel}_{turno_sel}", None)
                st.rerun()
        with c_a:
            if st.button(lbl_a, key=f"a_{eid}", use_container_width=True):
                st.session_state[sk(eid)] = "A"
                guardar_un_estado(eid, fecha_sel, turno_sel, "Ausente Injustificado")
                st.session_state.pop(f"asist_{grado_sel}_{fecha_sel}_{turno_sel}", None)
                st.rerun()
        with c_j:
            if st.button(lbl_j, key=f"j_{eid}", use_container_width=True):
                st.session_state[sk(eid)] = "J"
                guardar_un_estado(eid, fecha_sel, turno_sel, "Ausente Justificado")
                st.session_state.pop(f"asist_{grado_sel}_{fecha_sel}_{turno_sel}", None)
                st.rerun()
        with c_rep:
            if st.button("📋", key=f"rep_{eid}", use_container_width=True):
                cur = st.session_state.get("reporte_eid")
                st.session_state["reporte_eid"] = None if cur == eid else eid
                st.rerun()

        if st.session_state.get("reporte_eid") == eid:
            df_rep = get_reporte_estudiante(eid)
            if df_rep.empty:
                st.info("Sin registros aún.")
            else:
                presentes    = (df_rep["estado"] == "Presente").sum()
                ausentes     = (df_rep["estado"] == "Ausente Injustificado").sum()
                justificados = (df_rep["estado"] == "Ausente Justificado").sum()
                total_rep    = len(df_rep)
                pct = round(presentes / total_rep * 100) if total_rep > 0 else 0
                r1, r2, r3, r4 = st.columns(4)
                r1.metric("✅ Presentes",    presentes)
                r2.metric("❌ Ausentes",     ausentes)
                r3.metric("📝 Justificados", justificados)
                r4.metric("📊 Asistencia",   f"{pct}%")
                for _, r in df_rep.head(20).iterrows():
                    fstr  = pd.to_datetime(r["fecha"]).strftime("%d/%m/%Y")
                    trn   = r.get("turno", "")
                    badge = {"Presente": "🟢 PRESENTE",
                             "Ausente Injustificado": "🔴 AUSENTE",
                             "Ausente Justificado":   "🟡 JUSTIFICADO"}.get(r["estado"], r["estado"])
                    st.markdown(
                        f'<div style="display:flex;justify-content:space-between;'
                        f'padding:6px 12px;border-radius:6px;margin-bottom:3px;'
                        f'background:rgba(128,128,128,0.07);">'
                        f'<span>{fstr} <small style="opacity:.5">{trn}</small></span>'
                        f'<span>{badge}</span></div>',
                        unsafe_allow_html=True)
            st.divider()


def pagina_resumen():
    st.header("📊 Resumen por Grado")
    institucion = st.session_state.get("cfg_nombre", "Institución Educativa")

    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        grado_sel = st.selectbox("Grado", TODOS_LOS_GRADOS, key="resumen_grado")
    with col2:
        fecha_ini = st.date_input("Desde", value=date.today().replace(day=1), key="res_ini")
    with col3:
        fecha_fin = st.date_input("Hasta", value=date.today(), key="res_fin")

    df = get_resumen_grado(grado_sel)
    if df.empty or df["total_dias"].sum() == 0:
        st.info(f"📭 Aún no hay registros para **{grado_sel}**.")
        return

    total_reg = df["total_dias"].sum()
    total_p   = df["presentes"].sum()
    total_inj = df["inj"].sum()
    total_just= df["just"].sum()
    pct = round((total_p/total_reg)*100, 1) if total_reg else 0

    col1, col2 = st.columns(2)
    col1.metric("👥 Estudiantes", len(df))
    col2.metric("✅ % Asistencia", f"{pct}%")
    col3, col4 = st.columns(2)
    col3.metric("🔴 F. Injustificadas", total_inj)
    col4.metric("🟡 F. Justificadas", total_just)

    st.divider()
    col_pie, col_bar = st.columns(2)
    with col_pie:
        st.subheader("Distribución")
        fig = px.pie(
            names=["Presentes","Inj.","Just."],
            values=[total_p, total_inj, total_just],
            color_discrete_sequence=["#2ecc71","#e74c3c","#f39c12"], hole=0.4,
        )
        fig.update_traces(textposition="inside", textinfo="percent+label")
        fig.update_layout(showlegend=False, margin=dict(t=10,b=10))
        st.plotly_chart(fig, use_container_width=True)

    with col_bar:
        st.subheader("Por Estudiante")
        df_p = df.copy()
        df_p["% Asistencia"] = (df_p["presentes"]/df_p["total_dias"]*100).round(1)
        fig2 = px.bar(df_p, x="nombre", y="% Asistencia", color="% Asistencia",
                      color_continuous_scale=["#e74c3c","#f39c12","#2ecc71"], range_color=[0,100])
        fig2.update_layout(xaxis_tickangle=-35, coloraxis_showscale=False, margin=dict(t=10,b=80))
        fig2.add_hline(y=75, line_dash="dash", line_color="red", annotation_text="Mínimo 75%")
        st.plotly_chart(fig2, use_container_width=True)

    st.subheader("📋 Detalle")
    df_t = df.rename(columns={"nombre":"Nombre","presentes":"Presentes",
                               "inj":"F. Injustificadas","just":"F. Justificadas","total_dias":"Días"})
    df_t["% Asistencia"] = (df_t["Presentes"]/df_t["Días"]*100).round(1).astype(str)+"%"
    st.dataframe(df_t, use_container_width=True, hide_index=True)

    st.divider()
    st.subheader("📥 Exportar a Excel")
    periodo = st.radio("Período", ["Día","Semana","Mes","Personalizado"], horizontal=True)
    hoy = date.today()
    if periodo == "Día":
        f_ini, f_fin = hoy, hoy
    elif periodo == "Semana":
        f_ini = hoy - timedelta(days=hoy.weekday()); f_fin = hoy
    elif periodo == "Mes":
        f_ini = hoy.replace(day=1); f_fin = hoy
    else:
        f_ini, f_fin = fecha_ini, fecha_fin

    if st.button("📊 Generar Excel", type="primary"):
        excel_bytes = generar_excel_resumen(grado_sel, f_ini, f_fin, institucion)
        st.download_button(
            "⬇️ Descargar Excel", data=excel_bytes,
            file_name=f"asistencia_{grado_sel.replace('°','').replace(' ','_')}_{f_ini}_{f_fin}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


def pagina_alertas():
    st.header("🚨 Alertas de Faltas Consecutivas")
    st.caption(f"Estudiantes con **{UMBRAL_FALTAS_CONSECUTIVAS} o más faltas consecutivas**.")

    filtro = st.selectbox("Filtrar por Grado", ["Todos los grados"]+TODOS_LOS_GRADOS, key="alerta_grado")
    df = detectar_faltas_consecutivas(None if filtro=="Todos los grados" else filtro)

    if df.empty:
        st.success("✅ ¡Sin alertas!")
        return

    st.error(f"⚠️ **{len(df)} estudiante(s)** requieren atención:")
    st.divider()
    for _, row in df.iterrows():
        c1, c2 = st.columns([3,2])
        with c1:
            st.markdown(f"### 👤 {row['nombre']}")
            st.markdown(f"📚 **Grado:** {row['grado']}")
            st.markdown(f"📅 **Faltas consecutivas:** `{row['faltas_consecutivas']} días`")
            st.markdown(f"🗓️ **Desde:** {row['desde']}")
        with c2:
            st.markdown("### 📞 Contacto")
            st.info(f"**{row['contacto']}**")
        st.divider()

    csv = df.to_csv(index=False).encode("utf-8")
    st.download_button("📥 Descargar CSV", data=csv,
                       file_name=f"alertas_{date.today()}.csv", mime="text/csv")


def pagina_gestion():
    st.header("🎓 Gestión de Estudiantes")
    tabs = st.tabs(["➕ Agregar","📥 Importar PDF","📊 Importar Excel Asistencia","✏️ Editar / Mover","📞 Contactos","🗑️ Eliminar"])

    with tabs[0]:
        with st.form("form_agregar"):
            nombre   = st.text_input("Nombre y apellido")
            ci       = st.text_input("CI (Cédula de Identidad)")
            grado_s  = st.selectbox("Grado", TODOS_LOS_GRADOS)
            contacto = st.text_input("Contacto padre/tutor (opcional)")
            if st.form_submit_button("Agregar Estudiante", type="primary"):
                if nombre.strip():
                    agregar_estudiante(nombre, ci, grado_s, contacto)
                    st.success(f"✅ **{nombre}** agregado a {grado_s}.")
                else:
                    st.error("⚠️ El nombre no puede estar vacío.")

    with tabs[1]:
        st.subheader("📥 Importar listado desde PDF")
        st.info("El PDF debe tener tabla con columnas **Nombre** y **CI**.")
        grado_import = st.selectbox("Grado destino", TODOS_LOS_GRADOS, key="grado_import")
        pdf_file = st.file_uploader("Subir PDF", type=["pdf"])
        if pdf_file:
            try:
                df_prev = extraer_alumnos_pdf(pdf_file.read())
                if df_prev.empty:
                    st.error("❌ No se encontraron alumnos. Verificá que el PDF tenga tabla con Nombre y CI.")
                else:
                    st.success(f"✅ **{len(df_prev)} alumnos** encontrados. Verificá:")
                    st.dataframe(df_prev, use_container_width=True, hide_index=True)
                    if st.button("📥 Confirmar Importación", type="primary"):
                        conn = get_conn()
                        with conn.cursor() as cur:
                            cur.execute("SELECT id FROM grados WHERE nombre=%s", (grado_import,))
                            gr = cur.fetchone()
                            if gr:
                                count = 0
                                for _, r in df_prev.iterrows():
                                    if r["nombre"].strip():
                                        cur.execute(
                                            "INSERT INTO estudiantes (nombre,ci,grado_id) VALUES (%s,%s,%s)",
                                            (r["nombre"].strip(), r.get("ci",""), gr[0]),
                                        )
                                        count += 1
                                conn.commit()
                                st.success(f"✅ {count} alumnos importados a **{grado_import}**.")
            except Exception as e:
                st.error(f"❌ Error al leer el PDF: {e}")

    with tabs[2]:
        st.subheader("📊 Importar historial de asistencia desde Excel")
        st.info("""**Formato esperado:** igual al reporte exportado de la app.
        - Fila 1: encabezados — Documento | Nombre Completo | 1/3 | 3/3 | ...
        - Las fechas se interpretan como del año actual
        - Valores válidos: P, A, J (el guión `-` se ignora)
        """)

        col_g, col_t = st.columns(2)
        grado_xl = col_g.selectbox("Grado destino", TODOS_LOS_GRADOS, key="xl_grado")
        turno_xl = col_t.radio("Turno", TURNOS, horizontal=True, key="xl_turno")
        xl_file  = st.file_uploader("Subir Excel (.xlsx)", type=["xlsx"], key="xl_upload")

        if xl_file:
            try:
                import openpyxl as _oxl
                from datetime import date as _date
                wb_xl = _oxl.load_workbook(io.BytesIO(xl_file.read()))
                ws_xl = wb_xl.active
                rows_xl = list(ws_xl.iter_rows(values_only=True))
                header  = rows_xl[0]

                # Detectar columnas de fecha (col 3 en adelante, hasta las de totales)
                año = _date.today().year
                fecha_cols = []
                for ci, h in enumerate(header):
                    if h and str(h).strip() and "/" in str(h):
                        try:
                            parts = str(h).strip().split("/")
                            if len(parts) == 2:
                                d, m = int(parts[0]), int(parts[1])
                                fecha_cols.append((ci, _date(año, m, d)))
                        except Exception:
                            pass

                if not fecha_cols:
                    st.error("❌ No se encontraron columnas de fecha en el encabezado.")
                else:
                    # Preview
                    preview = []
                    for row in rows_xl[1:]:
                        ci_val = str(row[0] or "").strip()
                        nombre = str(row[1] or "").strip()
                        if not nombre:
                            continue
                        for ci, fecha in fecha_cols:
                            val = str(row[ci] or "").strip().upper() if ci < len(row) else ""
                            if val in ("P","A","J"):
                                preview.append({"Nombre": nombre, "Fecha": fecha.strftime("%d/%m/%Y"),
                                                "Estado": {"P":"Presente","A":"Ausente Injustificado","J":"Ausente Justificado"}[val]})
                    
                    if not preview:
                        st.warning("⚠️ No se encontraron registros válidos (P/A/J).")
                    else:
                        df_prev_xl = pd.DataFrame(preview)
                        st.success(f"✅ **{len(preview)} registros** encontrados en {len(fecha_cols)} fechas.")
                        st.dataframe(df_prev_xl.head(20), use_container_width=True, hide_index=True)
                        if len(preview) > 20:
                            st.caption(f"...y {len(preview)-20} registros más")

                        if st.button("📥 Importar a la BD", type="primary", key="btn_import_xl"):
                            ESTADO_MAP = {"P":"Presente","A":"Ausente Injustificado","J":"Ausente Justificado"}
                            # Obtener IDs de estudiantes por nombre o CI
                            df_ests = get_estudiantes_por_grado(grado_xl)
                            nombre_id = {r["nombre"].strip().upper(): int(r["id"]) for _, r in df_ests.iterrows()}
                            ci_id     = {str(r["ci"]).strip(): int(r["id"]) for _, r in df_ests.iterrows() if r.get("ci")}

                            ok, skip = 0, 0
                            registros_bulk = []
                            for row in rows_xl[1:]:
                                ci_val = str(row[0] or "").strip()
                                nombre = str(row[1] or "").strip().upper()
                                if not nombre:
                                    continue
                                est_id = ci_id.get(ci_val) or nombre_id.get(nombre)
                                if not est_id:
                                    skip += 1
                                    continue
                                for ci, fecha in fecha_cols:
                                    val = str(row[ci] or "").strip().upper() if ci < len(row) else ""
                                    if val in ("P","A","J"):
                                        registros_bulk.append((est_id, fecha, turno_xl, ESTADO_MAP[val]))
                                        ok += 1

                            # Insertar todo en una sola query bulk
                            if registros_bulk:
                                conn = get_conn()
                                with conn.cursor() as cur:
                                    psycopg2.extras.execute_values(cur, """
                                        INSERT INTO asistencia (estudiante_id, fecha, turno, estado)
                                        VALUES %s
                                        ON CONFLICT (estudiante_id, fecha, turno)
                                        DO UPDATE SET estado=EXCLUDED.estado
                                    """, registros_bulk, page_size=200)
                                conn.commit()

                            # Limpiar cachés
                            for k in [k for k in st.session_state if k.startswith(("asist_","resumen_","rep_","grados_init"))]:
                                st.session_state.pop(k, None)
                            st.session_state.discard if hasattr(st.session_state, "discard") else None

                            st.success(f"✅ **{ok}** registros importados.")
                            if skip:
                                st.warning(f"⚠️ {skip} alumnos no encontrados en {grado_xl} — verificá que estén cargados.")
                            if err:
                                st.error(f"❌ {err} errores al guardar.")

            except Exception as e:
                st.error(f"❌ Error al leer el archivo: {e}")

    with tabs[3]:
        st.subheader("✏️ Editar datos o mover de grado")
        grado_ver = st.selectbox("Ver estudiantes de", TODOS_LOS_GRADOS, key="editar_grado")
        df_est = get_estudiantes_por_grado(grado_ver)
        if df_est.empty:
            st.info(f"No hay estudiantes en {grado_ver}.")
        else:
            alumno_sel = st.selectbox("Seleccionar alumno", df_est["nombre"].tolist(), key="alumno_editar")
            ar = df_est[df_est["nombre"]==alumno_sel].iloc[0]
            with st.form("form_editar"):
                nn = st.text_input("Nombre", value=ar["nombre"])
                nc = st.text_input("CI", value=ar.get("ci","") or "")
                ng = st.selectbox("Grado", TODOS_LOS_GRADOS, index=TODOS_LOS_GRADOS.index(grado_ver))
                nco= st.text_input("Contacto", value=ar.get("contacto","") or "")
                if st.form_submit_button("💾 Guardar Cambios", type="primary"):
                    try:
                        actualizar_estudiante(ar["id"], nn, nc, ng, nco)
                        st.success(f"✅ Datos de **{nn}** actualizados.")
                        st.rerun()
                    except Exception as ex:
                        st.error(f"❌ Error al guardar: {ex}")

    with tabs[4]:
        st.subheader("📞 Contactos de estudiantes")
        st.caption("Agregá o editá el número de contacto. Tocá el botón para abrir WhatsApp.")
        grado_cont = st.selectbox("Grado", TODOS_LOS_GRADOS, key="cont_grado")
        df_cont = get_estudiantes_por_grado(grado_cont)
        if df_cont.empty:
            st.info(f"No hay estudiantes en {grado_cont}.")
        else:
            for _, row in df_cont.iterrows():
                contacto_actual = str(row.get("contacto", "") or "")
                c1, c2, c3 = st.columns([3, 2, 1])
                c1.write(f"**{row['nombre']}**")

                # Campo editable para el número
                nuevo_num = c2.text_input(
                    "Número",
                    value=contacto_actual,
                    key=f"cont_{row['id']}",
                    placeholder="0981-xxx-xxx",
                    label_visibility="collapsed",
                )

                # Guardar si cambió
                if nuevo_num != contacto_actual:
                    conn = get_conn()
                    with conn.cursor() as cur:
                        cur.execute("UPDATE estudiantes SET contacto=%s WHERE id=%s", (nuevo_num.strip(), row["id"]))
                        conn.commit()

                # Botón WhatsApp — abre chat directo
                if contacto_actual:
                    # Limpiar número: solo dígitos, agregar código Paraguay si empieza con 0
                    num_limpio = "".join(filter(str.isdigit, contacto_actual))
                    if num_limpio.startswith("0"):
                        num_limpio = "595" + num_limpio[1:]
                    wa_url = f"https://wa.me/{num_limpio}"
                    c3.markdown(
                        f'<a href="{wa_url}" target="_blank" style="text-decoration:none;font-size:22px;">💬</a>',
                        unsafe_allow_html=True,
                    )
                else:
                    c3.write("—")

                st.markdown("<hr style='margin:4px 0;opacity:0.1'>", unsafe_allow_html=True)

    with tabs[4]:
        st.subheader("🗑️ Eliminar estudiante")
        grado_del = st.selectbox("Grado", TODOS_LOS_GRADOS, key="del_grado")
        df_del = get_estudiantes_por_grado(grado_del)
        if df_del.empty:
            st.info(f"No hay estudiantes en {grado_del}.")
        else:
            for _, row in df_del.iterrows():
                c1, c2, c3 = st.columns([3,2,1])
                c1.write(f"👤 {row['nombre']}")
                c2.write(f"🪪 {row.get('ci','') or '—'}")
                if c3.button("🗑️", key=f"del_{row['id']}"):
                    eliminar_estudiante(row["id"])
                    st.rerun()


def pagina_configuracion():
    st.header("⚙️ Configuración Institucional")
    nombre_actual = get_config("institucion_nombre")
    logo_actual   = get_config("institucion_logo")

    with st.form("form_config"):
        st.subheader("🏫 Nombre de la institución")
        nuevo_nombre = st.text_input("Nombre", value=nombre_actual)

        st.subheader("🖼️ Logo institucional")
        logo_file = st.file_uploader("Subir logo (PNG/JPG)", type=["png","jpg","jpeg"])
        if logo_actual:
            st.markdown("**Logo actual:**")
            st.markdown(
                f'<img src="data:image/png;base64,{logo_actual}" style="max-height:80px;border-radius:8px;">',
                unsafe_allow_html=True,
            )

        if st.form_submit_button("💾 Guardar", type="primary"):
            set_config("institucion_nombre", nuevo_nombre.strip())
            if logo_file:
                set_config("institucion_logo", base64.b64encode(logo_file.read()).decode("utf-8"))
            # Limpiar caché de configuración para que se recargue
            st.session_state.pop("cfg_nombre", None)
            st.session_state.pop("cfg_logo", None)
            st.success("✅ Configuración guardada.")
            st.rerun()


def pagina_reportes():
    st.header("📨 Reportes para compartir")

    col1, col2 = st.columns([2, 1])
    with col1:
        fecha_sel = st.date_input("Fecha del reporte", value=date.today(), key="rep_fecha")
    with col2:
        grado_ausentes = st.selectbox("Grado (ausentes)", ["Todos"] + TODOS_LOS_GRADOS, key="rep_grado")

    st.markdown("---")

    # ── Reporte 1: Justificados del día (todos los grados) ──
    st.subheader("📝 Justificados del día")
    st.caption("Alumnos con ausencia justificada — para enviar a profesores")

    df_just = run_df("""
        SELECT e.nombre, g.nombre as grado
        FROM asistencia a
        JOIN estudiantes e ON a.estudiante_id = e.id
        JOIN grados g ON e.grado_id = g.id
        WHERE a.fecha = %s AND a.estado = 'Ausente Justificado'
        ORDER BY g.nombre, e.nombre
    """, (fecha_sel,))

    if df_just.empty:
        st.info(f"✅ Sin justificados el {fecha_sel.strftime('%d/%m/%Y')}.")
    else:
        # Construir texto para copiar
        lineas = [f"📋 *JUSTIFICADOS — {fecha_sel.strftime('%d/%m/%Y')}*", ""]
        grado_actual = None
        for _, r in df_just.iterrows():
            if r["grado"] != grado_actual:
                if grado_actual is not None:
                    lineas.append("")
                grado_actual = r["grado"]
                lineas.append(f"*{grado_actual}*")
            lineas.append(f"  • {r['nombre']}")
        lineas.append("")
        lineas.append(f"_Total: {len(df_just)} alumno(s)_")
        texto_just = "\n".join(lineas)

        st.text_area(
            "Tocá el texto, seleccioná todo (Ctrl+A / ⌘+A) y copiá:",
            value=texto_just,
            height=max(150, len(lineas) * 22),
            key="txt_justificados",
        )
        st.caption(f"📊 {len(df_just)} justificado(s) en {df_just['grado'].nunique()} grado(s)")

    st.markdown("---")

    # ── Reporte 2: Ausentes sin justificar por grado ──
    st.subheader("🔴 Ausentes sin justificar")
    st.caption("Alumnos ausentes injustificados — para notificar por grado")

    query_params = (fecha_sel,)
    filtro_grado = ""
    if grado_ausentes != "Todos":
        filtro_grado = "AND g.nombre = %s"
        query_params = (fecha_sel, grado_ausentes)

    df_aus = run_df(f"""
        SELECT e.nombre, g.nombre as grado
        FROM asistencia a
        JOIN estudiantes e ON a.estudiante_id = e.id
        JOIN grados g ON e.grado_id = g.id
        WHERE a.fecha = %s AND a.estado = 'Ausente Injustificado'
        {filtro_grado}
        ORDER BY g.nombre, e.nombre
    """, query_params)

    if df_aus.empty:
        st.info(f"✅ Sin ausentes injustificados el {fecha_sel.strftime('%d/%m/%Y')}" +
                (f" en {grado_ausentes}." if grado_ausentes != "Todos" else "."))
    else:
        lineas2 = [f"🔴 *AUSENTES — {fecha_sel.strftime('%d/%m/%Y')}*"]
        if grado_ausentes != "Todos":
            lineas2[0] += f" — *{grado_ausentes}*"
        lineas2.append("")
        grado_actual = None
        for _, r in df_aus.iterrows():
            if grado_ausentes == "Todos" and r["grado"] != grado_actual:
                if grado_actual is not None:
                    lineas2.append("")
                grado_actual = r["grado"]
                lineas2.append(f"*{grado_actual}*")
            lineas2.append(f"  • {r['nombre']}")
        lineas2.append("")
        lineas2.append(f"_Total: {len(df_aus)} alumno(s)_")
        texto_aus = "\n".join(lineas2)

        st.text_area(
            "Tocá el texto, seleccioná todo (Ctrl+A / ⌘+A) y copiá:",
            value=texto_aus,
            height=max(150, len(lineas2) * 22),
            key="txt_ausentes",
        )
        st.caption(f"📊 {len(df_aus)} ausente(s) injustificado(s)")



# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    st.set_page_config(
        page_title="Asistencia Escolar — MEC Paraguay",
        page_icon="🏫",
        layout="wide",
    )

    # ── LOGIN ──
    if not verificar_login():
        pagina_login()
        st.stop()

    inject_css()

    # ── Init BD solo una vez por sesión ──
    if not st.session_state.get("db_inicializada_v2"):
        try:
            init_db()
            seed_mock_data()
            st.session_state["db_inicializada_v2"] = True
            st.session_state.pop("db_inicializada", None)
        except Exception as e:
            st.error(f"❌ No se pudo conectar a la base de datos.\n\n`{e}`")
            st.stop()

    # Config en caché — solo consulta si no está en session_state
    if "cfg_nombre" not in st.session_state:
        try:
            st.session_state["cfg_nombre"] = get_config("institucion_nombre")
            st.session_state["cfg_logo"]   = get_config("institucion_logo")
        except Exception:
            st.session_state["cfg_nombre"] = "Institución Educativa"
            st.session_state["cfg_logo"]   = ""
    institucion_nombre = st.session_state["cfg_nombre"]
    logo_b64           = st.session_state["cfg_logo"]

    # CSS para botones de navegación del sidebar
    st.markdown("""
    <style>
    /* Botones de navegación: ancho completo, alineados a la izquierda, tamaño cómodo */
    [data-testid="stSidebar"] .stButton > button {
        width: 100% !important;
        text-align: left !important;
        justify-content: flex-start !important;
        padding: 12px 16px !important;
        font-size: 15px !important;
        font-weight: 500 !important;
        border-radius: 10px !important;
        border: none !important;
        background: transparent !important;
        margin-bottom: 2px !important;
        min-height: 50px !important;
    }
    [data-testid="stSidebar"] .stButton > button:hover {
        background: rgba(255,255,255,0.08) !important;
    }
    .nav-activo button {
        background: rgba(46,204,113,0.15) !important;
        border-left: 3px solid #2ecc71 !important;
        color: #2ecc71 !important;
        font-weight: 700 !important;
    }
    </style>
    """, unsafe_allow_html=True)

    PAGINAS = [
        ("📋 Pasar Lista",          "lista"),
        ("📨 Reportes",             "reportes"),
        ("📊 Resumen por Grado",    "resumen"),
        ("🚨 Alertas de Faltas",    "alertas"),
        ("🎓 Gestión de Estudiantes","gestion"),
        ("⚙️ Configuración",        "config"),
    ]

    if "pagina_sel" not in st.session_state:
        st.session_state["pagina_sel"] = "lista"

    with st.sidebar:
        if logo_b64:
            st.markdown(
                f'<img src="data:image/png;base64,{logo_b64}" style="max-height:72px;border-radius:8px;margin-bottom:6px;">',
                unsafe_allow_html=True,
            )
        else:
            st.image("https://upload.wikimedia.org/wikipedia/commons/2/27/Flag_of_Paraguay.svg", width=70)

        st.markdown(f"### {institucion_nombre}")
        st.caption("MEC Paraguay · Tercer Ciclo & Nivel Medio")
        st.divider()

        for label, key in PAGINAS:
            activo = st.session_state["pagina_sel"] == key
            css_class = "nav-activo" if activo else ""
            st.markdown(f'<div class="{css_class}">', unsafe_allow_html=True)
            if st.button(label, key=f"nav_{key}"):
                st.session_state["pagina_sel"] = key
                st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)

        st.divider()
        st.caption(f"📅 Hoy: {date.today().strftime('%d/%m/%Y')}")
        # Métricas cacheadas — se actualizan cada 5 minutos
        import time
        ahora = time.time()
        if "metrics_ts" not in st.session_state or ahora - st.session_state["metrics_ts"] > 300:
            try:
                st.session_state["n_est"] = run_df("SELECT COUNT(*) as c FROM estudiantes")["c"][0]
                st.session_state["n_reg"] = run_df("SELECT COUNT(*) as c FROM asistencia")["c"][0]
                st.session_state["metrics_ts"] = ahora
            except Exception:
                st.session_state.setdefault("n_est", "—")
                st.session_state.setdefault("n_reg", "—")
        st.metric("Total Estudiantes", st.session_state.get("n_est", "—"))
        st.metric("Registros de Asistencia", st.session_state.get("n_reg", "—"))
        st.divider()
        if st.button("🚪 Cerrar sesión", key="nav_logout"):
            st.session_state["autenticado"] = False
            st.query_params.clear()
            st.rerun()

    pagina = st.session_state["pagina_sel"]
    # Autoguardar lista si el usuario navega a otra página
    if pagina != "lista" and st.session_state.get("lista_prev_key"):
        prev_key = st.session_state.get("lista_prev_key")
        prev_parts = prev_key.split("|") if prev_key else []
        if len(prev_parts) == 3:
            prev_grado, prev_fecha_str, prev_turno = prev_parts
            try:
                prev_fecha = date.fromisoformat(prev_fecha_str)
                prev_ck    = f"asist_{prev_grado}_{prev_fecha_str}_{prev_turno}"
                prev_df    = st.session_state.get(prev_ck)
                if prev_df is not None and not prev_df.empty:
                    def nav_sk(eid): return f"est_{prev_grado}_{prev_fecha_str}_{prev_turno}_{eid}"
                    autoguardar_lista(prev_df, prev_fecha, prev_turno, nav_sk)
            except Exception:
                pass

    if pagina == "lista":
        pagina_pasar_lista()
    elif pagina == "reportes":
        pagina_reportes()
    elif pagina == "resumen":
        pagina_resumen()
    elif pagina == "alertas":
        pagina_alertas()
    elif pagina == "gestion":
        pagina_gestion()
    elif pagina == "config":
        pagina_configuracion()


if __name__ == "__main__":
    main()
